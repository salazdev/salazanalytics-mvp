import streamlit as st
import pandas as pd
from datetime import datetime, date
import base64
import io
import json

INSTRUCCIONES_IMPORTAR = """**Columnas requeridas:**\n\n- **Fecha** — formato DD/MM/AAAA (ej: 15/01/2026)\n- **Tipo** — exactamente: Ingreso o Gasto\n- **Categoria** — ver hoja Categorias en la plantilla\n- **Descripcion** — texto libre describiendo el movimiento\n- **Valor Base** — numero entero sin IVA (ej: 3500000)\n- **IVA %** — solo 0, 5 o 19\n\nLas filas 2-4 son ejemplos, puedes borrarlas. Puedes importar hasta 200 movimientos por archivo."""

# ─────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────

CATEGORIAS_INGRESO = [
    "Ventas de productos",
    "Prestación de servicios",
    "Honorarios",
    "Arrendamientos",
    "Intereses y rendimientos",
    "Otros ingresos",
]

CATEGORIAS_GASTO = [
    "Nómina y salarios",
    "Arriendo",
    "Servicios públicos",
    "Internet y telecomunicaciones",
    "Publicidad y marketing",
    "Contabilidad y revisoría",
    "Software y suscripciones",
    "Transporte y logística",
    "Compra de mercancía",
    "Equipos y herramientas",
    "Gastos bancarios",
    "Impuestos y tasas",
    "Capacitación",
    "Otros gastos",
]

MESES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
         "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

# Tarifa SIMPLE 2024 — servicios profesionales (grupo 3)
TARIFAS_SIMPLE = {
    "0-6.000":    2.0,
    "6.000-15.000": 2.8,
    "15.000-30.000": 5.5,
    "30.000-80.000": 7.0,
    ">80.000":    8.5,
}

# ─────────────────────────────────────────────
# INIT STATE
# ─────────────────────────────────────────────

def _init():
    """Sin estado local — los datos viven en SQLite."""
    pass

def _nit() -> str:
    return st.session_state.get("nit", "DEMO")

def _movimientos() -> list:
    from pathlib import Path
    import importlib.util, sys
    base = Path(__file__).parent
    spec = importlib.util.spec_from_file_location("db", base / "_db.py")
    mod  = importlib.util.module_from_spec(spec)
    sys.modules["db"] = mod
    spec.loader.exec_module(mod)
    return mod.movimientos_listar(_nit())

def _db_mod():
    from pathlib import Path
    import importlib.util, sys
    base = Path(__file__).parent
    spec = importlib.util.spec_from_file_location("db", base / "_db.py")
    mod  = importlib.util.module_from_spec(spec)
    sys.modules["db"] = mod
    spec.loader.exec_module(mod)
    return mod

def _df() -> pd.DataFrame:
    mvs = _movimientos()
    if not mvs:
        return pd.DataFrame(columns=["id","fecha","tipo","categoria","descripcion","valor","iva","valor_iva","total"])
    df = pd.DataFrame(mvs)
    df["fecha"] = pd.to_datetime(df["fecha"])
    df = df.sort_values("fecha", ascending=False).reset_index(drop=True)
    return df

# ─────────────────────────────────────────────
# EXPORTAR EXCEL
# ─────────────────────────────────────────────

def _exportar_excel() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    COLOR_DARK  = "0D1B2A"
    COLOR_CYAN  = "00C2FF"
    COLOR_GREEN = "00A86B"
    COLOR_RED   = "E53935"
    COLOR_GRAY  = "7B9BB5"
    COLOR_LIGHT = "E8F4FD"

    wb = Workbook()

    # ── Hoja 1: Movimientos ──
    ws1 = wb.active
    ws1.title = "Movimientos"

    headers = ["Fecha", "Tipo", "Categoría", "Descripción", "Valor Base", "IVA %", "Valor IVA", "Total"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color=COLOR_CYAN, name="Arial", size=10)
        cell.fill = PatternFill("solid", start_color=COLOR_DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    df = _df()
    for row_idx, row in df.iterrows():
        r = row_idx + 2
        ws1.cell(r, 1, str(row["fecha"])[:10])
        ws1.cell(r, 2, row["tipo"])
        ws1.cell(r, 3, row["categoria"])
        ws1.cell(r, 4, row["descripcion"])
        ws1.cell(r, 5, row["valor"])
        ws1.cell(r, 6, row["iva"])
        ws1.cell(r, 7, row["valor_iva"])
        ws1.cell(r, 8, row["total"])

        # color fila por tipo
        fill_color = "E8F5E9" if row["tipo"] == "Ingreso" else "FFEBEE"
        for col in range(1, 9):
            cell = ws1.cell(r, col)
            cell.fill = PatternFill("solid", start_color=fill_color)
            cell.font = Font(name="Arial", size=9)
            if col in [5, 7, 8]:
                cell.number_format = '$#,##0'

    # anchos
    for col, w in zip(range(1, 9), [12, 10, 22, 35, 14, 8, 12, 14]):
        ws1.column_dimensions[get_column_letter(col)].width = w

    # ── Hoja 2: Resumen mensual ──
    ws2 = wb.create_sheet("Resumen Mensual")
    df2 = _df().copy()
    if not df2.empty:
        df2["mes"] = df2["fecha"].dt.month
        df2["año"] = df2["fecha"].dt.year

        headers2 = ["Mes", "Ingresos", "Gastos", "Utilidad", "IVA Cobrado", "IVA Pagado", "IVA Neto"]
        for col, h in enumerate(headers2, 1):
            cell = ws2.cell(1, col, h)
            cell.font = Font(bold=True, color=COLOR_CYAN, name="Arial", size=10)
            cell.fill = PatternFill("solid", start_color=COLOR_DARK)
            cell.alignment = Alignment(horizontal="center")

        grupos = df2.groupby(["año", "mes"])
        r = 2
        for (año, mes), g in sorted(grupos):
            ingresos  = g[g["tipo"]=="Ingreso"]["total"].sum()
            gastos    = g[g["tipo"]=="Gasto"]["total"].sum()
            iva_cobrado = g[g["tipo"]=="Ingreso"]["valor_iva"].sum()
            iva_pagado  = g[g["tipo"]=="Gasto"]["valor_iva"].sum()

            ws2.cell(r, 1, f"{MESES[mes-1]} {año}")
            ws2.cell(r, 2, ingresos)
            ws2.cell(r, 3, gastos)
            ws2.cell(r, 4, f"=B{r}-C{r}")
            ws2.cell(r, 5, iva_cobrado)
            ws2.cell(r, 6, iva_pagado)
            ws2.cell(r, 7, f"=E{r}-F{r}")

            for col in range(2, 8):
                ws2.cell(r, col).number_format = '$#,##0'
            r += 1

        # Totales
        last = r - 1
        ws2.cell(r, 1, "TOTAL").font = Font(bold=True, name="Arial")
        for col in range(2, 8):
            ws2.cell(r, col, f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{last})")
            ws2.cell(r, col).number_format = '$#,##0'
            ws2.cell(r, col).font = Font(bold=True, name="Arial")

        for col in range(1, 8):
            ws2.column_dimensions[get_column_letter(col)].width = 18

    # ── Hoja 3: IVA SIMPLE ──
    ws3 = wb.create_sheet("IVA y SIMPLE")
    ws3["A1"] = "Liquidación IVA y SIMPLE"
    ws3["A1"].font = Font(bold=True, color=COLOR_CYAN, name="Arial", size=12)

    if not df2.empty:
        total_ingresos  = df2[df2["tipo"]=="Ingreso"]["total"].sum()
        total_gastos    = df2[df2["tipo"]=="Gasto"]["total"].sum()
        iva_cobrado     = df2[df2["tipo"]=="Ingreso"]["valor_iva"].sum()
        iva_pagado      = df2[df2["tipo"]=="Gasto"]["valor_iva"].sum()
        iva_neto        = iva_cobrado - iva_pagado
        ingresos_base   = df2[df2["tipo"]=="Ingreso"]["valor"].sum()

        # Estimar SIMPLE (tarifa 2.0% base para simplificar — el usuario ajusta)
        simple_estimado = ingresos_base * 0.02

        datos = [
            ("", ""),
            ("INGRESOS TOTALES", total_ingresos),
            ("GASTOS TOTALES", total_gastos),
            ("UTILIDAD BRUTA", total_ingresos - total_gastos),
            ("", ""),
            ("IVA COBRADO (ventas)", iva_cobrado),
            ("IVA PAGADO (compras)", iva_pagado),
            ("IVA A PAGAR / FAVOR", iva_neto),
            ("", ""),
            ("BASE INGRESOS SIMPLE", ingresos_base),
            ("TARIFA SIMPLE aprox. 2%", simple_estimado),
            ("* Verifique tarifa con su contador", ""),
        ]
        for i, (label, valor) in enumerate(datos, 3):
            ws3[f"A{i}"] = label
            if valor != "":
                ws3[f"B{i}"] = valor
                ws3[f"B{i}"].number_format = '$#,##0'
            if label and not label.startswith("*"):
                ws3[f"A{i}"].font = Font(name="Arial", size=10)
                ws3[f"B{i}"].font = Font(name="Arial", size=10, bold=True)

        ws3.column_dimensions["A"].width = 30
        ws3.column_dimensions["B"].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ─────────────────────────────────────────────
# AGENTE IA — registra movimiento desde texto
# ─────────────────────────────────────────────

def _agente_registrar(texto: str) -> dict:
    import anthropic, re

    system = """Eres el asistente contable de SalazAnalytics para empresas colombianas en Régimen Simple.
Extrae del texto un movimiento contable y devuelve SOLO JSON válido sin markdown:

{
  "tipo": "Ingreso" o "Gasto",
  "categoria": string (una de las categorías colombianas típicas),
  "descripcion": string corto,
  "valor": número entero en pesos colombianos (sin IVA),
  "iva": 0 o 5 o 19,
  "mensaje": "confirmación amigable en español"
}

Si el valor mencionado incluye IVA, descuéntalo para obtener la base.
Si no se menciona IVA, usa 0."""

    client = anthropic.Anthropic()
    resp = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=512,
        system=system,
        messages=[{"role": "user", "content": texto}]
    )
    raw = re.sub(r"```json|```", "", resp.content[0].text).strip()
    return json.loads(raw)


# ─────────────────────────────────────────────
# PLANTILLA Y IMPORTADOR EXCEL
# ─────────────────────────────────────────────

CATEGORIAS_TODAS = [
    "Ventas de productos", "Prestación de servicios", "Honorarios",
    "Arrendamientos", "Intereses y rendimientos", "Otros ingresos",
    "Nómina y salarios", "Arriendo", "Servicios públicos",
    "Internet y telecomunicaciones", "Publicidad y marketing",
    "Contabilidad y revisoría", "Software y suscripciones",
    "Transporte y logística", "Compra de mercancía", "Equipos y herramientas",
    "Gastos bancarios", "Impuestos y tasas", "Capacitación", "Otros gastos",
]

def _generar_plantilla_excel() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.data_validation import DataValidation
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    COLOR_DARK  = "0D1B2A"
    COLOR_CYAN  = "00C2FF"
    COLOR_GRAY  = "7B9BB5"
    COLOR_LIGHT = "E8F4FD"
    COLOR_GREEN = "E8F5E9"
    COLOR_RED   = "FFEBEE"

    # ── Encabezados ──
    headers = [
        ("fecha",       "Fecha (DD/MM/AAAA)", 16),
        ("tipo",        "Tipo (Ingreso/Gasto)", 18),
        ("categoria",   "Categoría", 30),
        ("descripcion", "Descripción", 40),
        ("valor",       "Valor Base (sin IVA)", 16),
        ("iva",         "IVA % (0-5-19)", 14),
    ]

    for col, (_, label, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font      = Font(bold=True, color=COLOR_CYAN, name="Arial", size=10)
        cell.fill      = PatternFill("solid", start_color=COLOR_DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 36

    # ── Ejemplos (filas 2-4) ──
    ejemplos = [
        ("15/01/2026", "Ingreso", "Honorarios",         "Consultoría financiera enero - Cliente ABC",  3500000, 19),
        ("20/01/2026", "Gasto",   "Nómina y salarios",  "Nómina enero empleados",                    12250905,  0),
        ("31/01/2026", "Gasto",   "Arriendo",           "Arriendo oficina enero",                     1800000,  0),
    ]
    for row, datos in enumerate(ejemplos, 2):
        for col, val in enumerate(datos, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = Font(name="Arial", size=9, italic=True, color="888888")
            cell.fill = PatternFill("solid", start_color="F5F5F5")
            cell.alignment = Alignment(horizontal="center" if col in [1,2,5,6] else "left")

    # ── Filas vacías para datos (5 a 204) ──
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(5, 205):
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.font      = Font(name="Arial", size=9)
            cell.border    = border
            cell.alignment = Alignment(horizontal="center" if col in [1,2,5,6] else "left",
                                       vertical="center")
            # Colorear fondo por tipo (col 2)
            if col == 2:
                cell.fill = PatternFill("solid", start_color="FAFAFA")

    # ── Validación: Tipo ──
    dv_tipo = DataValidation(type="list", formula1='"Ingreso,Gasto"',
                              allow_blank=False, showDropDown=False)
    dv_tipo.error      = "Escribe Ingreso o Gasto"
    dv_tipo.errorTitle = "Valor inválido"
    ws.add_data_validation(dv_tipo)
    dv_tipo.add(f"B5:B204")

    # ── Validación: IVA ──
    dv_iva = DataValidation(type="list", formula1='"0,5,19"',
                             allow_blank=False, showDropDown=False)
    dv_iva.error      = "Solo se permite 0, 5 o 19"
    dv_iva.errorTitle = "IVA inválido"
    ws.add_data_validation(dv_iva)
    dv_iva.add("F5:F204")

    # ── Validación: Valor positivo ──
    dv_valor = DataValidation(type="whole", operator="greaterThan",
                               formula1="0", allow_blank=True)
    dv_valor.error      = "El valor debe ser mayor a 0"
    dv_valor.errorTitle = "Valor inválido"
    ws.add_data_validation(dv_valor)
    dv_valor.add("E5:E204")

    # ── Hoja de instrucciones ──
    ws2 = wb.create_sheet("Instrucciones")
    instrucciones = [
        ("INSTRUCCIONES DE USO", None, True, "0D1B2A", "00C2FF"),
        ("", None, False, None, None),
        ("1. FECHA", "Formato DD/MM/AAAA. Ej: 15/01/2026", False, None, None),
        ("2. TIPO", "Escribe exactamente: Ingreso o Gasto (con mayúscula)", False, None, None),
        ("3. CATEGORÍA", "Copia una de las categorías de la hoja 'Categorías'", False, None, None),
        ("4. DESCRIPCIÓN", "Describe brevemente el movimiento (máx 100 caracteres)", False, None, None),
        ("5. VALOR BASE", "El valor SIN IVA en pesos colombianos. Solo números enteros.", False, None, None),
        ("6. IVA %", "Escribe 0, 5 o 19. Si no aplica IVA escribe 0.", False, None, None),
        ("", None, False, None, None),
        ("IMPORTANTE", "No modifiques los encabezados de la fila 1.", False, None, None),
        ("", "Las filas 2, 3 y 4 son ejemplos — puedes borrarlos.", False, None, None),
        ("", "Puedes agregar hasta 200 movimientos por importación.", False, None, None),
        ("", "El sistema calculará automáticamente el IVA y el total.", False, None, None),
    ]
    for row, (label, valor, bold, bg, fg) in enumerate(instrucciones, 1):
        c1 = ws2.cell(row=row, column=1, value=label)
        c2 = ws2.cell(row=row, column=2, value=valor)
        if bold:
            c1.font = Font(bold=True, name="Arial", size=12,
                           color=fg or "000000")
            c1.fill = PatternFill("solid", start_color=bg or "FFFFFF")
        else:
            c1.font = Font(bold=bool(label), name="Arial", size=10)
            c2.font = Font(name="Arial", size=10)
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 60

    # ── Hoja de categorías válidas ──
    ws3 = wb.create_sheet("Categorías")
    ws3.cell(1, 1, "INGRESOS").font = Font(bold=True, color="00C2FF", name="Arial")
    ws3.cell(1, 2, "GASTOS").font   = Font(bold=True, color="E53935", name="Arial")
    ingresos = [c for c in CATEGORIAS_TODAS if c in [
        "Ventas de productos","Prestación de servicios","Honorarios",
        "Arrendamientos","Intereses y rendimientos","Otros ingresos"]]
    gastos   = [c for c in CATEGORIAS_TODAS if c not in ingresos]
    for i, cat in enumerate(ingresos, 2):
        ws3.cell(i, 1, cat).font = Font(name="Arial", size=9)
    for i, cat in enumerate(gastos, 2):
        ws3.cell(i, 2, cat).font = Font(name="Arial", size=9)
    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 35

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _importar_desde_excel(archivo, nit_empresa: str) -> tuple[int, list[str]]:
    """
    Lee el Excel del cliente y guarda los movimientos en SQLite.
    Retorna (cantidad_importados, lista_de_errores)
    """
    from datetime import datetime

    try:
        df = pd.read_excel(archivo, sheet_name="Movimientos", header=0, skiprows=0)
    except Exception as e:
        return 0, [f"No se pudo leer el archivo: {e}"]

    # Normalizar nombres de columnas
    df.columns = [str(c).strip().lower().split("\n")[0].split("(")[0].strip()
                  for c in df.columns]

    # Mapeo flexible de columnas
    col_map = {
        "fecha":       ["fecha"],
        "tipo":        ["tipo"],
        "categoria":   ["categoría", "categoria"],
        "descripcion": ["descripción", "descripcion"],
        "valor":       ["valor base", "valor"],
        "iva":         ["iva %", "iva"],
    }
    rename = {}
    for target, posibles in col_map.items():
        for p in posibles:
            if p in df.columns:
                rename[p] = target
                break

    df = df.rename(columns=rename)

    # Validar columnas requeridas
    requeridas = ["fecha", "tipo", "categoria", "descripcion", "valor", "iva"]
    faltantes = [r for r in requeridas if r not in df.columns]
    if faltantes:
        return 0, [f"Columnas faltantes: {', '.join(faltantes)}. "
                   f"Descarga la plantilla oficial y úsala."]

    # Filtrar filas vacías y ejemplos (filas 2-4 del Excel = índices 0-2)
    df = df.dropna(subset=["fecha", "tipo", "valor"])
    df = df[df["tipo"].astype(str).str.strip().isin(["Ingreso", "Gasto"])]

    if df.empty:
        return 0, ["El archivo no tiene movimientos válidos. "
                   "Asegúrate de que la columna Tipo diga 'Ingreso' o 'Gasto'."]

    db = _db_mod()
    importados = 0
    errores    = []

    for idx, row in df.iterrows():
        fila = idx + 2  # número de fila en Excel (1-indexed + encabezado)
        try:
            # Fecha
            fecha_raw = row["fecha"]
            if pd.isna(fecha_raw):
                errores.append(f"Fila {fila}: fecha vacía — omitida.")
                continue
            if isinstance(fecha_raw, str):
                for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"]:
                    try:
                        fecha = datetime.strptime(fecha_raw.strip(), fmt).date()
                        break
                    except:
                        fecha = None
                if not fecha:
                    errores.append(f"Fila {fila}: fecha '{fecha_raw}' inválida — usa DD/MM/AAAA.")
                    continue
            else:
                try:
                    fecha = pd.to_datetime(fecha_raw).date()
                except:
                    errores.append(f"Fila {fila}: fecha inválida — omitida.")
                    continue

            # Tipo
            tipo = str(row["tipo"]).strip()
            if tipo not in ["Ingreso", "Gasto"]:
                errores.append(f"Fila {fila}: tipo '{tipo}' inválido — debe ser Ingreso o Gasto.")
                continue

            # Categoría
            categoria = str(row.get("categoria", "Otros gastos" if tipo == "Gasto" else "Otros ingresos")).strip()
            if categoria not in CATEGORIAS_TODAS:
                categoria = "Otros gastos" if tipo == "Gasto" else "Otros ingresos"

            # Descripción
            descripcion = str(row.get("descripcion", "Sin descripción")).strip()[:200]
            if not descripcion or descripcion == "nan":
                descripcion = "Sin descripción"

            # Valor
            try:
                valor = float(str(row["valor"]).replace(",", "").replace("$", "").strip())
                if valor <= 0:
                    errores.append(f"Fila {fila}: valor {valor} no es válido — debe ser > 0.")
                    continue
            except:
                errores.append(f"Fila {fila}: valor '{row['valor']}' no es un número.")
                continue

            # IVA
            try:
                iva = int(float(str(row.get("iva", 0)).replace("%", "").strip()))
                if iva not in [0, 5, 19]:
                    iva = 0
            except:
                iva = 0

            valor_iva = round(valor * iva / 100)
            total     = round(valor + valor_iva)

            db.movimiento_crear(nit_empresa, str(fecha), tipo, categoria,
                                descripcion, valor, iva, valor_iva, total)
            importados += 1

        except Exception as e:
            errores.append(f"Fila {fila}: error inesperado — {e}")

    return importados, errores


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def show():
    _init()

    st.markdown("## 📒 Contabilidad")
    st.markdown(
        "<p style='color:#7B9BB5'>Ingresos, gastos, IVA y estimado SIMPLE — "
        "todo en un solo lugar.</p>",
        unsafe_allow_html=True
    )

    tab_agente, tab_manual, tab_dashboard, tab_iva, tab_oblig, tab_importar = st.tabs([
        "🤖 Registrar con IA", "✏️ Registro Manual", "📊 Dashboard", "🧾 IVA y SIMPLE", "⚖️ Obligaciones", "📥 Importar Excel"
    ])

    # ══════════════════════════════════════════
    # TAB AGENTE IA
    # ══════════════════════════════════════════
    with tab_agente:
        st.markdown("### Registra un movimiento en lenguaje natural")
        st.markdown(
            "<p style='color:#7B9BB5'>Cuéntale al agente qué pasó y él lo clasifica.</p>",
            unsafe_allow_html=True
        )

        with st.expander("💡 Ejemplos"):
            st.markdown("""
- *"Pagué arriendo de oficina $1.200.000"*
- *"Cobré honorarios a cliente ABC por $3.500.000 más IVA"*
- *"Compré papelería $85.000 sin IVA"*
- *"Recibí pago de servicio de contabilidad $800.000 con IVA del 19%"*
- *"Nómina de mayo $2.800.000"*
            """)

        texto = st.text_area(
            "Describe el movimiento:",
            placeholder="Ej: Pagué el internet de mayo $120.000...",
            height=100,
            key="agente_contab_texto"
        )

        if st.button("🤖 Registrar con IA", type="primary", key="btn_agente_contab"):
            if not texto.strip():
                st.warning("Escribe el movimiento primero.")
            else:
                with st.spinner("Clasificando movimiento..."):
                    try:
                        r = _agente_registrar(texto)
                        valor_iva = round(r["valor"] * r["iva"] / 100)
                        total = r["valor"] + valor_iva
                        mov = {
                            "fecha":       str(date.today()),
                            "tipo":        r["tipo"],
                            "categoria":   r["categoria"],
                            "descripcion": r["descripcion"],
                            "valor":       r["valor"],
                            "iva":         r["iva"],
                            "valor_iva":   valor_iva,
                            "total":       total,
                        }
                        db = _db_mod()
                        db.movimiento_crear(_nit(), mov["fecha"], mov["tipo"],
                            mov["categoria"], mov["descripcion"], mov["valor"],
                            mov["iva"], mov["valor_iva"], mov["total"])
                        tipo_color = "🟢" if r["tipo"] == "Ingreso" else "🔴"
                        st.success(
                            f"{tipo_color} {r['mensaje']}  \n"
                            f"**{r['tipo']}** · {r['categoria']} · "
                            f"Base: ${r['valor']:,.0f} · IVA: ${valor_iva:,.0f} · "
                            f"Total: ${total:,.0f}"
                        )
                    except Exception as e:
                        st.error(f"Error del agente: {e}")

    # ══════════════════════════════════════════
    # TAB REGISTRO MANUAL
    # ══════════════════════════════════════════
    with tab_manual:
        st.markdown("### Nuevo movimiento")

        c1, c2 = st.columns(2)
        with c1:
            tipo = st.radio("Tipo", ["Ingreso", "Gasto"], horizontal=True, key="m_tipo")
        with c2:
            fecha_mov = st.date_input("Fecha", value=date.today(), key="m_fecha_mov")

        categorias = CATEGORIAS_INGRESO if tipo == "Ingreso" else CATEGORIAS_GASTO
        categoria = st.selectbox("Categoría", categorias, key="m_categoria")
        descripcion = st.text_input("Descripción", key="m_descripcion",
                                    placeholder="Ej: Honorarios cliente XYZ")

        c1, c2, c3 = st.columns(3)
        with c1:
            valor = st.number_input("Valor base (sin IVA)", min_value=0, step=1000,
                                    key="m_valor", value=0)
        with c2:
            iva_pct = st.selectbox("IVA %", [0, 5, 19], key="m_iva")
        with c3:
            valor_iva = valor * iva_pct / 100
            total_mov = valor + valor_iva
            st.metric("Total con IVA", f"${total_mov:,.0f}")

        if st.button("➕ Guardar movimiento", type="primary", key="btn_guardar_mov"):
            if not descripcion:
                st.warning("Agrega una descripción.")
            elif valor <= 0:
                st.warning("El valor debe ser mayor a 0.")
            else:
                db = _db_mod()
                db.movimiento_crear(_nit(), str(fecha_mov), tipo, categoria,
                    descripcion, valor, iva_pct, round(valor_iva), round(total_mov))
                st.success(f"✅ Movimiento guardado — ${total_mov:,.0f}")

        # ── Tabla de movimientos ──
        df = _df()
        if not df.empty:
            st.divider()
            st.markdown("### Movimientos registrados")

            # Filtros
            c1, c2, c3 = st.columns(3)
            with c1:
                filtro_tipo = st.selectbox("Filtrar por tipo", ["Todos","Ingreso","Gasto"],
                                           key="filtro_tipo")
            with c2:
                meses_disponibles = sorted(df["fecha"].dt.month.unique())
                filtro_mes = st.selectbox(
                    "Mes",
                    ["Todos"] + [MESES[m-1] for m in meses_disponibles],
                    key="filtro_mes"
                )
            with c3:
                filtro_cat = st.selectbox(
                    "Categoría",
                    ["Todas"] + sorted(df["categoria"].unique().tolist()),
                    key="filtro_cat"
                )

            df_f = df.copy()
            if filtro_tipo != "Todos":
                df_f = df_f[df_f["tipo"] == filtro_tipo]
            if filtro_mes != "Todos":
                mes_num = MESES.index(filtro_mes) + 1
                df_f = df_f[df_f["fecha"].dt.month == mes_num]
            if filtro_cat != "Todas":
                df_f = df_f[df_f["categoria"] == filtro_cat]

            # Mostrar tabla
            df_show = df_f[["fecha","tipo","categoria","descripcion","valor","iva","valor_iva","total"]].copy()
            df_show["fecha"] = df_show["fecha"].dt.strftime("%d/%m/%Y")
            df_show.columns = ["Fecha","Tipo","Categoría","Descripción","Base","IVA%","Valor IVA","Total"]
            st.dataframe(df_show, use_container_width=True, height=350)

            # Botón eliminar último
            c1, c2 = st.columns([1, 3])
            with c1:
                if st.button("🗑️ Eliminar último", key="btn_del_ultimo"):
                    mvs = _movimientos()
                    if mvs:
                        db = _db_mod()
                        db.movimiento_eliminar(mvs[0]["id"], _nit())
                        st.rerun()
            with c2:
                if st.button("📥 Exportar a Excel", key="btn_exportar", type="primary"):
                    xlsx = _exportar_excel()
                    b64 = base64.b64encode(xlsx).decode()
                    mes_actual = MESES[date.today().month - 1]
                    href = (
                        f'<a href="data:application/vnd.openxmlformats-officedocument.'
                        f'spreadsheetml.sheet;base64,{b64}" '
                        f'download="contabilidad_{mes_actual}_{date.today().year}.xlsx" '
                        f'style="display:inline-block;background:#00C2FF;color:#0D1B2A;'
                        f'font-weight:700;padding:.6rem 1.2rem;border-radius:8px;'
                        f'text-decoration:none;">⬇️ Descargar Excel</a>'
                    )
                    st.markdown(href, unsafe_allow_html=True)

    # ══════════════════════════════════════════
    # TAB DASHBOARD
    # ══════════════════════════════════════════
    with tab_dashboard:
        df = _df()
        if df.empty:
            st.info("Registra movimientos para ver el dashboard.")
        else:
            import plotly.express as px
            import plotly.graph_objects as go

            total_ingresos = df[df["tipo"]=="Ingreso"]["total"].sum()
            total_gastos   = df[df["tipo"]=="Gasto"]["total"].sum()
            utilidad       = total_ingresos - total_gastos
            margen         = (utilidad / total_ingresos * 100) if total_ingresos > 0 else 0

            # Métricas principales
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("💰 Ingresos", f"${total_ingresos:,.0f}")
            c2.metric("💸 Gastos",   f"${total_gastos:,.0f}")
            color_util = "normal" if utilidad >= 0 else "inverse"
            c3.metric("📈 Utilidad", f"${utilidad:,.0f}", delta=f"{margen:.1f}% margen")
            c4.metric("📝 Movimientos", len(df))

            st.divider()

            # Gráfico mensual
            df2 = df.copy()
            df2["mes_num"] = df2["fecha"].dt.month
            df2["mes"]     = df2["fecha"].dt.month.apply(lambda m: MESES[m-1])
            df2["año"]     = df2["fecha"].dt.year

            resumen = df2.groupby(["año","mes_num","mes","tipo"])["total"].sum().reset_index()
            resumen = resumen.sort_values("mes_num")

            if not resumen.empty:
                fig = px.bar(
                    resumen, x="mes", y="total", color="tipo", barmode="group",
                    color_discrete_map={"Ingreso":"#00C2FF","Gasto":"#7B2FBE"},
                    title="Ingresos vs Gastos por mes",
                    labels={"total":"Valor ($)","mes":"Mes","tipo":"Tipo"},
                    template="plotly_dark"
                )
                fig.update_layout(
                    plot_bgcolor="#0D1B2A",
                    paper_bgcolor="#0D1B2A",
                    font_color="#E8F4FD",
                    title_font_color="#00C2FF"
                )
                st.plotly_chart(fig, use_container_width=True)

            c1, c2 = st.columns(2)

            # Donut gastos por categoría
            with c1:
                gastos_cat = df[df["tipo"]=="Gasto"].groupby("categoria")["total"].sum().reset_index()
                if not gastos_cat.empty:
                    fig2 = px.pie(
                        gastos_cat, values="total", names="categoria",
                        title="Distribución de gastos",
                        hole=0.5,
                        template="plotly_dark",
                        color_discrete_sequence=px.colors.sequential.Blues_r
                    )
                    fig2.update_layout(
                        plot_bgcolor="#0D1B2A", paper_bgcolor="#0D1B2A",
                        font_color="#E8F4FD", title_font_color="#00C2FF"
                    )
                    st.plotly_chart(fig2, use_container_width=True)

            # Top ingresos por categoría
            with c2:
                ing_cat = df[df["tipo"]=="Ingreso"].groupby("categoria")["total"].sum().reset_index()
                ing_cat = ing_cat.sort_values("total", ascending=True)
                if not ing_cat.empty:
                    fig3 = px.bar(
                        ing_cat, x="total", y="categoria", orientation="h",
                        title="Ingresos por categoría",
                        template="plotly_dark",
                        color_discrete_sequence=["#00C2FF"]
                    )
                    fig3.update_layout(
                        plot_bgcolor="#0D1B2A", paper_bgcolor="#0D1B2A",
                        font_color="#E8F4FD", title_font_color="#00C2FF"
                    )
                    st.plotly_chart(fig3, use_container_width=True)

    # ══════════════════════════════════════════
    # TAB IVA Y SIMPLE
    # ══════════════════════════════════════════
    with tab_iva:
        df = _df()
        if df.empty:
            st.info("Registra movimientos para ver la liquidación.")
        else:
            st.markdown("### 🧾 Liquidación IVA")

            # Filtro período
            años = sorted(df["fecha"].dt.year.unique(), reverse=True)
            c1, c2 = st.columns(2)
            with c1:
                año_sel = st.selectbox("Año", años, key="iva_año")
            with c2:
                periodo = st.selectbox(
                    "Período IVA",
                    ["Bimestre 1 (Ene-Feb)", "Bimestre 2 (Mar-Abr)",
                     "Bimestre 3 (May-Jun)", "Bimestre 4 (Jul-Ago)",
                     "Bimestre 5 (Sep-Oct)", "Bimestre 6 (Nov-Dic)",
                     "Cuatrimestre 1 (Ene-Abr)", "Cuatrimestre 2 (May-Ago)",
                     "Cuatrimestre 3 (Sep-Dic)", "Año completo"],
                    key="iva_periodo"
                )

            # Mapear período a meses
            mapa_periodos = {
                "Bimestre 1 (Ene-Feb)": [1,2],
                "Bimestre 2 (Mar-Abr)": [3,4],
                "Bimestre 3 (May-Jun)": [5,6],
                "Bimestre 4 (Jul-Ago)": [7,8],
                "Bimestre 5 (Sep-Oct)": [9,10],
                "Bimestre 6 (Nov-Dic)": [11,12],
                "Cuatrimestre 1 (Ene-Abr)": [1,2,3,4],
                "Cuatrimestre 2 (May-Ago)": [5,6,7,8],
                "Cuatrimestre 3 (Sep-Dic)": [9,10,11,12],
                "Año completo": list(range(1,13)),
            }
            meses_sel = mapa_periodos[periodo]

            df_p = df[
                (df["fecha"].dt.year == año_sel) &
                (df["fecha"].dt.month.isin(meses_sel))
            ]

            if df_p.empty:
                st.warning("No hay movimientos en ese período.")
            else:
                iva_cobrado = df_p[df_p["tipo"]=="Ingreso"]["valor_iva"].sum()
                iva_pagado  = df_p[df_p["tipo"]=="Gasto"]["valor_iva"].sum()
                iva_neto    = iva_cobrado - iva_pagado

                c1, c2, c3 = st.columns(3)
                c1.metric("IVA Cobrado (ventas)", f"${iva_cobrado:,.0f}")
                c2.metric("IVA Descontable (compras)", f"${iva_pagado:,.0f}")
                color = "inverse" if iva_neto > 0 else "normal"
                c3.metric(
                    "IVA a Pagar" if iva_neto > 0 else "Saldo a Favor",
                    f"${abs(iva_neto):,.0f}",
                )

                st.divider()
                st.markdown("### 📊 Estimado SIMPLE")
                st.markdown(
                    "<p style='color:#7B9BB5;font-size:.85rem'>"
                    "Estimación orientativa — confirma con tu contador.</p>",
                    unsafe_allow_html=True
                )

                ingresos_base = df_p[df_p["tipo"]=="Ingreso"]["valor"].sum()
                ingresos_uvt  = ingresos_base / 52_374  # UVT 2024

                # Determinar tarifa
                if ingresos_uvt <= 6000:
                    tarifa = 2.0
                    rango = "0 - 6.000 UVT"
                elif ingresos_uvt <= 15000:
                    tarifa = 2.8
                    rango = "6.000 - 15.000 UVT"
                elif ingresos_uvt <= 30000:
                    tarifa = 5.5
                    rango = "15.000 - 30.000 UVT"
                elif ingresos_uvt <= 80000:
                    tarifa = 7.0
                    rango = "30.000 - 80.000 UVT"
                else:
                    tarifa = 8.5
                    rango = "> 80.000 UVT"

                simple_estimado = ingresos_base * tarifa / 100

                c1, c2, c3, c4 = st.columns(4)
                c1.metric("Ingresos base período", f"${ingresos_base:,.0f}")
                c2.metric("Ingresos en UVT", f"{ingresos_uvt:,.1f}")
                c3.metric(f"Tarifa SIMPLE ({rango})", f"{tarifa}%")
                c4.metric("SIMPLE Estimado", f"${simple_estimado:,.0f}")

                st.info(
                    f"💡 Con ingresos de ${ingresos_base:,.0f} en este período, "
                    f"tu tarifa SIMPLE aplicable es del **{tarifa}%** "
                    f"(rango {rango}). SIMPLE estimado: **${simple_estimado:,.0f}**. "
                    f"Recuerda que SIMPLE ya incluye IVA, Renta, ICA y otras contribuciones."
                )

                # Exportar
                st.divider()
                if st.button("📥 Exportar informe completo Excel", type="primary", key="btn_exp_iva"):
                    xlsx = _exportar_excel()
                    b64 = base64.b64encode(xlsx).decode()
                    href = (
                        f'<a href="data:application/vnd.openxmlformats-officedocument.'
                        f'spreadsheetml.sheet;base64,{b64}" '
                        f'download="contabilidad_completa_{año_sel}.xlsx" '
                        f'style="display:inline-block;background:#00C2FF;color:#0D1B2A;'
                        f'font-weight:700;padding:.6rem 1.2rem;border-radius:8px;'
                        f'text-decoration:none;">⬇️ Descargar Excel Completo</a>'
                    )
                    st.markdown(href, unsafe_allow_html=True)

    # ══════════════════════════════════════════
    # TAB OBLIGACIONES TRIBUTARIAS
    # ══════════════════════════════════════════
    with tab_oblig:
        st.markdown("### ⚖️ Obligaciones Tributarias")
        st.markdown(
            "<p style='color:#7B9BB5'>Retención en la fuente, ICA, renta, "
            "prestaciones sociales y registros. Estimados para llevar a tu contador.</p>",
            unsafe_allow_html=True
        )

        UVT = 52_374  # UVT 2026

        sub_rete, sub_ica, sub_presta, sub_renta, sub_registros = st.tabs([
            "✂️ Retención en la Fuente",
            "🏙️ ICA Pereira",
            "👥 Prestaciones Sociales",
            "📋 Estimado Renta",
            "📜 Registros y Licencias",
        ])

        # ── RETENCIÓN EN LA FUENTE ──
        with sub_rete:
            st.markdown("#### Retención en la Fuente")
            st.markdown(
                "<p style='color:#7B9BB5;font-size:.85rem'>"
                "Calcula la retención que debes practicar al pagar a terceros.</p>",
                unsafe_allow_html=True
            )

            TARIFAS_RETE = {
                "Honorarios (profesionales)": {"tarifa": 11.0, "base_uvt": 0},
                "Servicios en general":        {"tarifa": 4.0,  "base_uvt": 4},
                "Servicios técnicos":          {"tarifa": 6.0,  "base_uvt": 4},
                "Compras / mercancía":         {"tarifa": 2.5,  "base_uvt": 27},
                "Arrendamientos (inmuebles)":  {"tarifa": 3.5,  "base_uvt": 0},
                "Arrendamientos (muebles)":    {"tarifa": 4.0,  "base_uvt": 0},
                "Transporte de carga":         {"tarifa": 1.0,  "base_uvt": 27},
                "Transporte de pasajeros":     {"tarifa": 3.5,  "base_uvt": 27},
            }

            c1, c2 = st.columns(2)
            with c1:
                concepto_rete = st.selectbox("Concepto del pago", list(TARIFAS_RETE.keys()), key="rete_concepto")
                valor_pago = st.number_input("Valor del pago ($)", min_value=0, step=100_000, key="rete_valor")
            with c2:
                declarante = st.radio("El beneficiario es", ["Declarante de renta", "No declarante"], key="rete_decl")
                mes_rete = st.selectbox("Mes del pago", MESES, key="rete_mes", index=date.today().month - 1)

            info = TARIFAS_RETE[concepto_rete]
            base_minima = info["base_uvt"] * UVT
            tarifa = info["tarifa"]
            # No declarantes pagan doble en servicios/honorarios
            if declarante == "No declarante" and concepto_rete in ["Honorarios (profesionales)", "Servicios en general", "Servicios técnicos"]:
                tarifa = min(tarifa * 2, 33.0)

            if valor_pago > 0 and valor_pago >= base_minima:
                retencion = round(valor_pago * tarifa / 100)
                neto_pagar = valor_pago - retencion
            else:
                retencion = 0
                neto_pagar = valor_pago

            st.divider()
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Valor bruto",    f"${valor_pago:,.0f}")
            c2.metric(f"Tarifa ({tarifa}%)", f"${retencion:,.0f}")
            c3.metric("Neto a pagar",   f"${neto_pagar:,.0f}")
            c4.metric("Base mínima",    f"${base_minima:,.0f}")

            if valor_pago > 0 and valor_pago < base_minima:
                st.info(f"⚠️ El pago (${valor_pago:,.0f}) no supera la base mínima de retención (${base_minima:,.0f}). **No aplica retención.**")
            elif retencion > 0:
                st.success(
                    f"✅ Debes retener **${retencion:,.0f}** al beneficiario y declararlo en el mes de {mes_rete}. "
                    f"Paga neto: **${neto_pagar:,.0f}**"
                )

            st.divider()
            st.markdown("##### Acumulado de retenciones del período")
            df = _df()
            if not df.empty:
                # Simular retenciones sobre pagos a terceros registrados
                gastos = df[df["tipo"] == "Gasto"].copy()
                gastos_hon = gastos[gastos["categoria"].isin(["Contabilidad y revisoría", "Otros gastos"])]
                total_base_hon = gastos_hon["valor"].sum()
                rete_hon = round(total_base_hon * 0.11)
                gastos_serv = gastos[gastos["categoria"].isin(["Software y suscripciones", "Publicidad y marketing", "Transporte y logística"])]
                total_base_serv = gastos_serv["valor"].sum()
                rete_serv = round(total_base_serv * 0.04)
                total_rete = rete_hon + rete_serv

                c1, c2, c3 = st.columns(3)
                c1.metric("Rete. honorarios estimada", f"${rete_hon:,.0f}")
                c2.metric("Rete. servicios estimada",  f"${rete_serv:,.0f}")
                c3.metric("Total retenciones período", f"${total_rete:,.0f}")
                st.caption("* Estimado basado en los gastos registrados. Valida con tu contador.")
            else:
                st.info("Registra movimientos para ver el acumulado de retenciones.")

        # ── ICA PEREIRA ──
        with sub_ica:
            st.markdown("#### Industria y Comercio — Pereira (Risaralda)")
            st.markdown(
                "<p style='color:#7B9BB5;font-size:.85rem'>"
                "Tarifa ICA Pereira 2024 · Declaración bimestral.</p>",
                unsafe_allow_html=True
            )

            TARIFAS_ICA = {
                "Servicios profesionales y consultoría": 8.0,
                "Actividades comerciales":               5.0,
                "Actividades industriales":              4.0,
                "Servicios financieros":                10.0,
                "Restaurantes y hoteles":                6.0,
                "Transporte":                            4.5,
                "Construcción":                          4.0,
                "Salud":                                 3.0,
                "Educación":                             2.0,
            }

            actividad_ica = st.selectbox(
                "Actividad económica principal",
                list(TARIFAS_ICA.keys()),
                key="ica_actividad"
            )
            tarifa_ica = TARIFAS_ICA[actividad_ica]

            st.markdown(f"**Tarifa aplicable:** {tarifa_ica} × 1.000 ({tarifa_ica/10:.1f}‰)")

            bimestres_ica = [
                "Bimestre 1 (Ene-Feb)", "Bimestre 2 (Mar-Abr)",
                "Bimestre 3 (May-Jun)", "Bimestre 4 (Jul-Ago)",
                "Bimestre 5 (Sep-Oct)", "Bimestre 6 (Nov-Dic)",
            ]
            mapa_bim = {
                "Bimestre 1 (Ene-Feb)": [1,2], "Bimestre 2 (Mar-Abr)": [3,4],
                "Bimestre 3 (May-Jun)": [5,6], "Bimestre 4 (Jul-Ago)": [7,8],
                "Bimestre 5 (Sep-Oct)": [9,10], "Bimestre 6 (Nov-Dic)": [11,12],
            }
            bim_sel = st.selectbox("Bimestre a declarar", bimestres_ica, key="ica_bim")
            meses_bim = mapa_bim[bim_sel]

            df = _df()
            if not df.empty:
                df_bim = df[
                    (df["tipo"] == "Ingreso") &
                    (df["fecha"].dt.month.isin(meses_bim))
                ]
                ingresos_bim = df_bim["valor"].sum()
                ica_bim = round(ingresos_bim * tarifa_ica / 1000)

                st.divider()
                c1, c2, c3 = st.columns(3)
                c1.metric("Ingresos del bimestre", f"${ingresos_bim:,.0f}")
                c2.metric(f"Tarifa ICA ({tarifa_ica}×1000)", f"{tarifa_ica/10:.1f}‰")
                c3.metric("ICA a pagar", f"${ica_bim:,.0f}")

                if ica_bim > 0:
                    st.success(
                        f"💡 Por el {bim_sel} debes declarar y pagar **${ica_bim:,.0f}** de ICA "
                        f"a la Alcaldía de Pereira. Plazo: últimos días del mes siguiente al bimestre."
                    )
            else:
                st.info("Registra ingresos para calcular el ICA.")

        # ── PRESTACIONES SOCIALES ──
        with sub_presta:
            st.markdown("#### Prestaciones Sociales y Carga Laboral")
            st.markdown(
                "<p style='color:#7B9BB5;font-size:.85rem'>"
                "Calcula cesantías, prima, vacaciones y parafiscales por empleado.</p>",
                unsafe_allow_html=True
            )

            st.markdown("##### Agregar empleado")
            c1, c2, c3 = st.columns(3)
            with c1:
                emp_nombre = st.text_input("Nombre empleado", key="emp_nombre")
                salario    = st.number_input("Salario básico ($)", min_value=0, step=100_000,
                                             value=1_750_905, key="emp_salario")
            with c2:
                aux_transporte = st.checkbox("Incluir auxilio de transporte ($249.095)", value=True, key="emp_aux")
                meses_trabajados = st.number_input("Meses trabajados (período)", min_value=1, max_value=12,
                                                    value=1, key="emp_meses")
            with c3:
                st.markdown("<br>", unsafe_allow_html=True)
                aux_val = 249_095 if aux_transporte else 0
                salario_total = salario + aux_val

                # Prestaciones (sobre salario sin auxilio transporte para algunos)
                cesantias      = round(salario * meses_trabajados / 12)
                int_cesantias  = round(cesantias * 0.12)
                prima          = round(salario_total * meses_trabajados / 12)
                vacaciones     = round(salario * meses_trabajados / 24)

                # Parafiscales (sobre salario)
                salud_emp      = round(salario * 0.04)   # empleado paga 4%
                pension_emp    = round(salario * 0.04)   # empleado paga 4%
                salud_emp_er   = round(salario * 0.085)  # empleador paga 8.5%
                pension_emp_er = round(salario * 0.12)   # empleador paga 12%
                arl            = round(salario * 0.00522) # clase I riesgo
                caja           = round(salario * 0.04)
                sena           = round(salario * 0.02)
                icbf           = round(salario * 0.03)

                carga_total = (cesantias + int_cesantias + prima + vacaciones +
                               salud_emp_er + pension_emp_er + arl + caja + sena + icbf)

                st.metric("Carga total empleador/mes", f"${carga_total + salario_total:,.0f}")

            if st.button("📊 Ver liquidación detallada", type="primary", key="btn_presta"):
                st.divider()
                st.markdown(f"##### Liquidación: {emp_nombre or 'Empleado'} — {meses_trabajados} mes(es)")

                c1, c2 = st.columns(2)
                with c1:
                    st.markdown("**Prestaciones sociales:**")
                    items_presta = [
                        ("Cesantías (8.33%)",          cesantias),
                        ("Intereses cesantías (1%)",    int_cesantias),
                        ("Prima de servicios (8.33%)",  prima),
                        ("Vacaciones (4.17%)",          vacaciones),
                    ]
                    for label, val in items_presta:
                        col_a, col_b = st.columns([3,1])
                        col_a.write(label)
                        col_b.write(f"**${val:,.0f}**")

                with c2:
                    st.markdown("**Seguridad social y parafiscales (empleador):**")
                    items_para = [
                        ("Salud empleador (8.5%)",   salud_emp_er),
                        ("Pensión empleador (12%)",  pension_emp_er),
                        ("ARL clase I (0.522%)",     arl),
                        ("Caja compensación (4%)",   caja),
                        ("SENA (2%)",                sena),
                        ("ICBF (3%)",                icbf),
                    ]
                    for label, val in items_para:
                        col_a, col_b = st.columns([3,1])
                        col_a.write(label)
                        col_b.write(f"**${val:,.0f}**")

                st.divider()
                col1, col2, col3, col4 = st.columns(4)
                col1.metric("Salario + auxilio", f"${salario_total:,.0f}")
                col2.metric("Prestaciones",      f"${cesantias+int_cesantias+prima+vacaciones:,.0f}")
                col3.metric("Parafiscales",      f"${salud_emp_er+pension_emp_er+arl+caja+sena+icbf:,.0f}")
                col4.metric("COSTO TOTAL/MES",   f"${carga_total + salario_total:,.0f}")

                pct_carga = (carga_total / salario * 100)
                st.info(f"💡 La carga prestacional y parafiscal representa el **{pct_carga:.1f}%** del salario básico. "
                        f"Por cada $1.000.000 de salario, el empleador paga aprox. ${pct_carga*10_000:,.0f} adicionales.")

        # ── ESTIMADO RENTA ──
        with sub_renta:
            st.markdown("#### Estimado Declaración de Renta")
            st.markdown(
                "<p style='color:#7B9BB5;font-size:.85rem'>"
                "Estimación orientativa para preparar la información con tu contador. "
                "No reemplaza la declaración oficial.</p>",
                unsafe_allow_html=True
            )

            df = _df()
            if df.empty:
                st.info("Registra movimientos para estimar la renta.")
            else:
                año_renta = st.selectbox("Año gravable", sorted(df["fecha"].dt.year.unique(), reverse=True), key="renta_año")
                df_año = df[df["fecha"].dt.year == año_renta]

                ingresos_brutos  = df_año[df_año["tipo"]=="Ingreso"]["valor"].sum()
                costos_gastos    = df_año[df_año["tipo"]=="Gasto"]["valor"].sum()
                renta_bruta      = ingresos_brutos - costos_gastos

                st.divider()
                st.markdown("##### Depuración básica de renta")

                c1, c2 = st.columns(2)
                with c1:
                    ded_gmf        = st.number_input("4×1000 (GMF) pagado", min_value=0, step=100_000, key="renta_gmf",
                                                      help="El 50% del GMF es deducible")
                    ded_medicina   = st.number_input("Medicina prepagada / seguros salud", min_value=0, step=100_000, key="renta_med")
                    ded_educacion  = st.number_input("Pagos educación (hasta 200 UVT)", min_value=0, step=100_000, key="renta_edu")
                with c2:
                    ded_vivienda   = st.number_input("Intereses crédito de vivienda", min_value=0, step=100_000, key="renta_viv")
                    ded_pension    = st.number_input("Aportes voluntarios pensión / AFC", min_value=0, step=100_000, key="renta_pen")
                    otros_ded      = st.number_input("Otras deducciones", min_value=0, step=100_000, key="renta_otros")

                total_deducciones = (ded_gmf * 0.5 + ded_medicina + ded_educacion +
                                     ded_vivienda + ded_pension + otros_ded)
                renta_liquida = max(renta_bruta - total_deducciones, 0)
                renta_uvt     = renta_liquida / UVT

                # Tabla de tarifas renta 2024 (Art. 241 ET)
                if renta_uvt <= 1090:
                    impuesto = 0
                    tarifa_renta = 0
                elif renta_uvt <= 1700:
                    impuesto = (renta_uvt - 1090) * 0.19 * UVT
                    tarifa_renta = 19
                elif renta_uvt <= 4100:
                    impuesto = ((renta_uvt - 1700) * 0.28 + 116) * UVT
                    tarifa_renta = 28
                elif renta_uvt <= 8670:
                    impuesto = ((renta_uvt - 4100) * 0.33 + 788) * UVT
                    tarifa_renta = 33
                elif renta_uvt <= 18970:
                    impuesto = ((renta_uvt - 8670) * 0.35 + 2296) * UVT
                    tarifa_renta = 35
                elif renta_uvt <= 31000:
                    impuesto = ((renta_uvt - 18970) * 0.37 + 5901) * UVT
                    tarifa_renta = 37
                else:
                    impuesto = ((renta_uvt - 31000) * 0.39 + 10352) * UVT
                    tarifa_renta = 39

                impuesto = round(impuesto)

                st.divider()
                c1, c2, c3, c4 = st.columns(4)
                c1.metric("Ingresos brutos",     f"${ingresos_brutos:,.0f}")
                c2.metric("Costos y gastos",      f"${costos_gastos:,.0f}")
                c3.metric("Deducciones",          f"${total_deducciones:,.0f}")
                c4.metric("Renta líquida",        f"${renta_liquida:,.0f}")

                st.divider()
                c1, c2, c3 = st.columns(3)
                c1.metric("Renta en UVT",         f"{renta_uvt:,.1f} UVT")
                c2.metric(f"Tarifa marginal",      f"{tarifa_renta}%")
                c3.metric("Impuesto estimado",     f"${impuesto:,.0f}")

                if impuesto == 0:
                    st.success("✅ Con estos datos no habría impuesto de renta a pagar. Confirma con tu contador.")
                else:
                    st.warning(
                        f"⚠️ Impuesto de renta estimado: **${impuesto:,.0f}** "
                        f"(tarifa marginal {tarifa_renta}%). Este es un estimado — tu contador "
                        f"puede optimizar deducciones y aplicar descuentos tributarios."
                    )

                st.caption(f"* Cálculo basado en tarifas Art. 241 ET. UVT 2024: ${UVT:,.0f}")

        # ── REGISTROS Y LICENCIAS ──
        with sub_registros:
            st.markdown("#### Registros, Licencias y Obligaciones Formales")
            st.markdown(
                "<p style='color:#7B9BB5;font-size:.85rem'>"
                "Calendario de obligaciones formales para empresas en Pereira.</p>",
                unsafe_allow_html=True
            )

            mes_actual = date.today().month
            año_actual = date.today().year

            obligaciones = [
                {
                    "obligacion": "Renovación Cámara de Comercio",
                    "periodicidad": "Anual",
                    "mes_limite": 3,
                    "descripcion": "Renovar matrícula mercantil antes del 31 de marzo",
                    "costo_aprox": "Según activos. Desde ~$150.000",
                    "entidad": "Cámara de Comercio de Pereira",
                },
                {
                    "obligacion": "Declaración ICA Bimestral",
                    "periodicidad": "Bimestral",
                    "mes_limite": None,
                    "descripcion": "Declarar y pagar ICA bimestral en la Alcaldía de Pereira",
                    "costo_aprox": "Según tarifa actividad",
                    "entidad": "Alcaldía de Pereira — Hacienda",
                },
                {
                    "obligacion": "Declaración IVA",
                    "periodicidad": "Bimestral/Cuatrimestral",
                    "mes_limite": None,
                    "descripcion": "Bimestral si ingresos > 92.000 UVT, cuatrimestral si menor",
                    "costo_aprox": "IVA cobrado - IVA descontable",
                    "entidad": "DIAN",
                },
                {
                    "obligacion": "Retención en la Fuente",
                    "periodicidad": "Mensual",
                    "mes_limite": None,
                    "descripcion": "Declarar y pagar retenciones practicadas el mes anterior",
                    "costo_aprox": "Según retenciones practicadas",
                    "entidad": "DIAN",
                },
                {
                    "obligacion": "SIMPLE / Anticipo bimestral",
                    "periodicidad": "Bimestral",
                    "mes_limite": None,
                    "descripcion": "Pagar anticipo bimestral del impuesto SIMPLE",
                    "costo_aprox": "Según tarifa e ingresos",
                    "entidad": "DIAN",
                },
                {
                    "obligacion": "Declaración de Renta",
                    "periodicidad": "Anual",
                    "mes_limite": 8,
                    "descripcion": "Personas naturales: agosto. Empresas: abril",
                    "costo_aprox": "Según impuesto liquidado",
                    "entidad": "DIAN",
                },
                {
                    "obligacion": "Información Exógena",
                    "periodicidad": "Anual",
                    "mes_limite": 4,
                    "descripcion": "Si superas topes DIAN, reportar transacciones del año anterior",
                    "costo_aprox": "Sin costo directo",
                    "entidad": "DIAN",
                },
                {
                    "obligacion": "Pago seguridad social (nómina)",
                    "periodicidad": "Mensual",
                    "mes_limite": None,
                    "descripcion": "Salud, pensión y ARL de empleados mediante PILA",
                    "costo_aprox": "24.522% del salario (empleador)",
                    "entidad": "Operadores PILA",
                },
            ]

            for ob in obligaciones:
                vence_pronto = (ob["mes_limite"] and ob["mes_limite"] == mes_actual)
                icon = "🔴" if vence_pronto else "🟡" if ob["periodicidad"] == "Mensual" else "🟢"
                with st.expander(f"{icon} {ob['obligacion']} — {ob['periodicidad']}"):
                    c1, c2 = st.columns(2)
                    c1.markdown(f"**Descripción:** {ob['descripcion']}")
                    c1.markdown(f"**Entidad:** {ob['entidad']}")
                    c2.markdown(f"**Periodicidad:** {ob['periodicidad']}")
                    c2.markdown(f"**Costo aprox:** {ob['costo_aprox']}")
                    if vence_pronto:
                        st.error(f"⚠️ Esta obligación vence este mes ({MESES[mes_actual-1]} {año_actual})")

            st.divider()
            st.info(
                "💡 **Consejo:** Lleva estas fechas en tu calendario y "
                "consulta con tu contador al menos 2 semanas antes de cada vencimiento. "
                "Las sanciones por extemporaneidad en Colombia pueden ser del 5% al 200% del impuesto."
            )

    # ══════════════════════════════════════════
    # TAB IMPORTAR EXCEL
    # ══════════════════════════════════════════
    with tab_importar:
        st.markdown("### 📥 Importar movimientos desde Excel")
        st.markdown(
            "<p style='color:#7B9BB5'>Carga todos tus movimientos de una vez "
            "usando la plantilla oficial de SalazAnalytics.</p>",
            unsafe_allow_html=True
        )

        # ── Paso 1: Descargar plantilla ──
        st.markdown("#### Paso 1 — Descarga la plantilla")
        st.markdown(
            "La plantilla tiene el formato exacto que necesita el sistema. "
            "Incluye instrucciones y las categorías válidas.",
            unsafe_allow_html=False
        )

        plantilla_bytes = _generar_plantilla_excel()
        b64_plantilla = base64.b64encode(plantilla_bytes).decode()
        href_plantilla = (
            f'<a href="data:application/vnd.openxmlformats-officedocument.'
            f'spreadsheetml.sheet;base64,{b64_plantilla}" '
            f'download="plantilla_movimientos_salazanalytics.xlsx" '
            f'style="display:inline-block;background:#00C2FF;color:#0D1B2A;'
            f'font-weight:700;padding:.7rem 1.5rem;border-radius:8px;'
            f'text-decoration:none;font-size:1rem;">'
            f'⬇️ Descargar Plantilla Excel</a>'
        )
        st.markdown(href_plantilla, unsafe_allow_html=True)

        st.divider()

        # ── Paso 2: Instrucciones rápidas ──
        st.markdown("#### Paso 2 — Llena la plantilla")
        with st.expander("Ver instrucciones de llenado"):
            st.markdown("""
**Columnas requeridas:**

| Columna | Formato | Ejemplo |
|---------|---------|---------|
| Fecha | DD/MM/AAAA | 15/01/2026 |
| Tipo | Ingreso o Gasto | Ingreso |
| Categoría | Ver hoja "Categorías" | Honorarios |
| Descripción | Texto libre | Consultoría enero |
| Valor Base | Número entero sin IVA | 3500000 |
| IVA % | 0, 5 o 19 | 19 |

**Consejos:**
- Las filas 2, 3 y 4 son ejemplos — puedes borrarlas o sobreescribirlas
- El valor debe ser la base sin IVA (el sistema calcula el IVA automáticamente)
- Si un movimiento no tiene IVA, escribe 0 en la columna IVA %
- Puedes importar hasta 200 movimientos por archivo
- Los duplicados no se detectan automáticamente — revisa antes de importar
            """)

        st.divider()

        # ── Paso 3: Subir y procesar ──
        st.markdown("#### Paso 3 — Sube el archivo lleno")

        archivo = st.file_uploader(
            "Selecciona el archivo Excel con tus movimientos",
            type=["xlsx", "xls"],
            key="importar_excel_file"
        )

        if archivo is not None:
            st.info(f"Archivo cargado: **{archivo.name}** "
                    f"({archivo.size / 1024:.1f} KB)")

            # Vista previa
            try:
                df_preview = pd.read_excel(archivo, sheet_name="Movimientos",
                                           header=0, nrows=8)
                st.markdown("**Vista previa (primeras filas):**")
                st.dataframe(df_preview, use_container_width=True)
                archivo.seek(0)
            except Exception as e:
                st.warning(f"No se pudo previsualizar: {e}")

            col1, col2 = st.columns([1, 3])
            with col1:
                confirmar = st.button("🚀 Importar movimientos",
                                      type="primary",
                                      key="btn_confirmar_importar")
            with col2:
                st.markdown(
                    "<p style='color:#7B9BB5;padding-top:.6rem'>"
                    "Los movimientos se agregarán a tu contabilidad actual.</p>",
                    unsafe_allow_html=True
                )

            if confirmar:
                with st.spinner("Procesando archivo..."):
                    archivo.seek(0)
                    importados, errores = _importar_desde_excel(
                        archivo, _nit()
                    )

                if importados > 0:
                    st.success(
                        f"✅ Se importaron **{importados} movimientos** "
                        f"exitosamente a tu contabilidad."
                    )

                if errores:
                    st.warning(f"Se encontraron {len(errores)} advertencias:")
                    with st.expander("Ver detalle de errores"):
                        for err in errores:
                            st.markdown(f"- {err}")

                if importados == 0 and not errores:
                    st.error("No se importó ningún movimiento. "
                             "Verifica que el archivo tenga datos válidos.")

        # ── Tip final ──
        st.divider()
        st.markdown(
            "<div style='background:#132030;border:1px solid #1a3a5c;"
            "border-radius:8px;padding:1rem;'>"
            "<p style='color:#00C2FF;font-weight:600;margin:0 0 .5rem'>💡 Tip para tus clientes</p>"
            "<p style='color:#7B9BB5;margin:0;font-size:.88rem'>"
            "Pídele al cliente que exporte su contabilidad actual a Excel "
            "(desde cualquier sistema — Siigo, Alegra, Helisa, o incluso un Excel manual) "
            "y que mapee las columnas a esta plantilla. "
            "Con esto puedes onboardear un cliente nuevo en menos de 10 minutos.</p>"
            "</div>",
            unsafe_allow_html=True
        )

    with tab_importar:
        st.markdown("### Importar movimientos desde Excel")
        st.markdown(
            "<p style='color:#7B9BB5'>Carga todos tus movimientos de una vez "
            "usando la plantilla oficial de SalazAnalytics.</p>",
            unsafe_allow_html=True
        )

        st.markdown("#### Paso 1 — Descarga la plantilla")
        plantilla_bytes = _generar_plantilla_excel()
        b64p = base64.b64encode(plantilla_bytes).decode()
        enlace = (
            '<a href="data:application/vnd.openxmlformats-officedocument.'
            'spreadsheetml.sheet;base64,' + b64p + '" '
            'download="plantilla_movimientos_salazanalytics.xlsx" '
            'style="display:inline-block;background:#00C2FF;color:#0D1B2A;'
            'font-weight:700;padding:.7rem 1.5rem;border-radius:8px;'
            'text-decoration:none;">Descargar Plantilla Excel</a>'
        )
        st.markdown(enlace, unsafe_allow_html=True)

        st.divider()
        st.markdown("#### Paso 2 — Llena la plantilla")
        with st.expander("Ver instrucciones"):
            st.markdown(INSTRUCCIONES_IMPORTAR)

        st.divider()
        st.markdown("#### Paso 3 — Sube el archivo lleno")

        archivo = st.file_uploader(
            "Selecciona el archivo Excel con tus movimientos",
            type=["xlsx","xls"],
            key="importar_excel_file"
        )

        if archivo is not None:
            kb = round(archivo.size / 1024, 1)
            st.info(f"Archivo cargado: {archivo.name} — {kb} KB")
            try:
                df_prev = pd.read_excel(archivo, sheet_name="Movimientos", header=0, nrows=6)
                st.markdown("**Vista previa:**")
                st.dataframe(df_prev, use_container_width=True)
                archivo.seek(0)
            except Exception as ex:
                st.warning(f"No se pudo previsualizar: {ex}")

            c1, c2 = st.columns([1, 3])
            with c1:
                confirmar = st.button("Importar movimientos", type="primary", key="btn_importar")
            with c2:
                st.markdown(
                    "<p style='color:#7B9BB5;padding-top:.6rem'>"
                    "Los movimientos se agregan a tu contabilidad actual.</p>",
                    unsafe_allow_html=True
                )

            if confirmar:
                with st.spinner("Procesando archivo..."):
                    archivo.seek(0)
                    importados, errores = _importar_desde_excel(archivo, _nit())
                if importados > 0:
                    st.success(f"Se importaron {importados} movimientos exitosamente.")
                if errores:
                    st.warning(f"Se encontraron {len(errores)} advertencias:")
                    with st.expander("Ver detalle"):
                        for err in errores:
                            st.markdown(f"- {err}")
                if importados == 0 and not errores:
                    st.error("No se importo ningun movimiento. Verifica el archivo.")

        st.divider()
        st.markdown(
            "<div style='background:#132030;border:1px solid #1a3a5c;"
            "border-radius:8px;padding:1rem;'>"
            "<p style='color:#00C2FF;font-weight:600;margin:0 0 .4rem'>Tip para tus clientes</p>"
            "<p style='color:#7B9BB5;margin:0;font-size:.88rem'>"
            "Pide al cliente que exporte su contabilidad a Excel "
            "(Siigo, Alegra, Helisa o Excel manual) y mapee las columnas a esta plantilla. "
            "Puedes incorporar un cliente nuevo en menos de 10 minutos.</p>"
            "</div>",
            unsafe_allow_html=True
        )
