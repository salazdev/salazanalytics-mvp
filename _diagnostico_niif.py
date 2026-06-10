"""
_diagnostico_niif.py — SalazAnalytics
Módulo: 🏛️ Diagnóstico NIIF Pymes — rediseño wizard simplificado
10 áreas · respuesta Sí / En proceso / No · semáforo en tiempo real
"""
import streamlit as st
import plotly.graph_objects as go
from datetime import datetime
from io import BytesIO

# ── Paleta ─────────────────────────────────────────────────────────────────
BG_DARK  = "#0D1B2A"; BG_CARD = "#132030"; BORDER = "#1a3a5c"
TEXT_SEC = "#7B9BB5"; ACCENT  = "#00C2FF"; DANGER = "#FF6B6B"
SUCCESS  = "#00FFB3"; WARN    = "#FFD93D"; PURPLE = "#7B2FBE"

SMMLV_2026 = 1_750_905

# ── 10 áreas de diagnóstico ────────────────────────────────────────────────
AREAS = [
    {
        "id": "politicas",
        "titulo": "📋 Políticas Contables",
        "pregunta": "¿Tu empresa tiene un manual de políticas contables escrito, aprobado y aplicado por el equipo?",
        "si":       "Manual documentado, socializado y actualizado en los últimos 2 años.",
        "proceso":  "Existe un borrador o se aplican políticas de manera informal.",
        "no":       "No hay manual ni políticas formalmente definidas.",
        "accion":   "Elaborar el manual de políticas contables adaptado a tu sector y marco NIIF.",
    },
    {
        "id": "ppe",
        "titulo": "🏭 Propiedades, Planta y Equipo",
        "pregunta": "¿Los activos fijos tienen vida útil, valor residual y depreciación calculados bajo NIIF (no la tabla fiscal)?",
        "si":       "Vidas útiles técnicas definidas, depreciación calculada y deterioro evaluado anualmente.",
        "proceso":  "Algunos activos están ajustados; otros aún usan tasas fiscales.",
        "no":       "Se deprecia con las tasas del Estatuto Tributario sin ajuste contable.",
        "accion":   "Revisar y ajustar vidas útiles, valores residuales y depreciación de todos los activos fijos.",
    },
    {
        "id": "inventarios",
        "titulo": "📦 Inventarios",
        "pregunta": "¿El inventario se valora al costo (FIFO o Promedio) y se evalúa si hay ítems por debajo del Valor Neto Realizable?",
        "si":       "Método de costeo definido, conteos físicos periódicos y VNR evaluado al cierre.",
        "proceso":  "Se tienen conteos pero no se evalúa el VNR ni se revisan obsolescencias.",
        "no":       "No hay método de costeo claro ni evaluación de inventario obsoleto.",
        "accion":   "Establecer método de costeo, rutina de conteos físicos y evaluación de VNR al cierre.",
    },
    {
        "id": "cartera",
        "titulo": "💳 Cartera y Provisiones",
        "pregunta": "¿Se analiza la cartera con criterios NIIF para estimar pérdidas crediticias (más allá de la provisión fiscal del 33%)?",
        "si":       "Modelo de deterioro basado en experiencia histórica y análisis por cliente.",
        "proceso":  "Se hace algún análisis pero no está documentado ni es sistemático.",
        "no":       "Solo se aplica la provisión fiscal del 33% para cartera mayor a 1 año.",
        "accion":   "Implementar matriz de provisión por antigüedad de cartera calibrada con datos históricos.",
    },
    {
        "id": "ingresos",
        "titulo": "💰 Reconocimiento de Ingresos",
        "pregunta": "¿Los ingresos se registran cuando se transfieren los riesgos al cliente (no cuando se cobra)?",
        "si":       "Ingresos causados al momento de entrega del bien o prestación del servicio.",
        "proceso":  "Se causa la mayoría pero hay casos de registro al cobro o al facturar.",
        "no":       "Los ingresos se registran cuando se recibe el pago (base caja).",
        "accion":   "Documentar política de reconocimiento de ingresos y ajustar el proceso contable.",
    },
    {
        "id": "pasivos",
        "titulo": "👥 Pasivos Laborales",
        "pregunta": "¿Las prestaciones sociales (cesantías, vacaciones, primas, intereses) se causan mensualmente?",
        "si":       "Causación mensual de todas las prestaciones; vacaciones provisionadas por días causados.",
        "proceso":  "Se causan algunas prestaciones mensualmente; otras solo en las fechas de pago.",
        "no":       "Las prestaciones se registran únicamente cuando se pagan (enero, junio, diciembre).",
        "accion":   "Implementar causación mensual de prestaciones. Usar el Simulador de Nómina de SalazAnalytics.",
    },
    {
        "id": "impuestos",
        "titulo": "🧾 Impuesto Diferido",
        "pregunta": "¿Se calcula el impuesto diferido por diferencias entre la base contable NIIF y la base fiscal?",
        "si":       "Impuesto diferido calculado, con conciliación patrimonio fiscal vs. contable.",
        "proceso":  "Se conoce el concepto pero el cálculo no está implementado formalmente.",
        "no":       "Solo se registra el impuesto corriente (lo que dice la declaración de renta).",
        "accion":   "Calcular diferencias temporarias y reconocer activos/pasivos por impuesto diferido.",
    },
    {
        "id": "eeff",
        "titulo": "📊 Estados Financieros",
        "pregunta": "¿Se presentan los 4 estados financieros completos bajo NIIF con período comparativo?",
        "si":       "Balance, P&G, Cambios en Patrimonio y Flujo de Efectivo con año anterior comparativo.",
        "proceso":  "Se preparan algunos estados pero sin todos los comparativos o sin el flujo de efectivo.",
        "no":       "Solo se prepara balance y estado de resultados, sin comparativos ni flujo de efectivo.",
        "accion":   "Completar el juego de estados financieros. Usar el módulo de Flujo Indirecto de SalazAnalytics.",
    },
    {
        "id": "notas",
        "titulo": "🔍 Notas y Revelaciones",
        "pregunta": "¿Las notas a los estados financieros detallan políticas, estimaciones, partes relacionadas y contingencias?",
        "si":       "Notas completas con todas las secciones requeridas por NIIF para Pymes.",
        "proceso":  "Hay notas básicas pero faltan revelaciones de partes relacionadas o contingencias.",
        "no":       "No se preparan notas o son genéricas sin información específica de la empresa.",
        "accion":   "Elaborar notas completas incluyendo partes relacionadas, contingencias y compromisos.",
    },
    {
        "id": "auditoria",
        "titulo": "🏛️ Auditoría y Entidades",
        "pregunta": "¿Los estados financieros han sido revisados por un auditor o presentados a Supersociedades/SIC bajo NIIF?",
        "si":       "Auditoría o revisión fiscal realizada bajo NIA; reportes a entidades al día.",
        "proceso":  "Se ha hecho alguna revisión interna pero no auditoría formal ni reporte a entidades.",
        "no":       "No se ha realizado auditoría ni se han presentado estados bajo NIIF a entidades de vigilancia.",
        "accion":   "Contratar auditoría bajo NIA y presentar estados financieros NIIF a las entidades correspondientes.",
    },
]

OPCIONES    = ["❓ Sin evaluar", "✅ Sí cumple", "⚠️ En proceso", "❌ No cumple"]
OP_COLORES  = {"✅ Sí cumple": SUCCESS, "⚠️ En proceso": WARN,
               "❌ No cumple": DANGER, "❓ Sin evaluar": TEXT_SEC}
OP_SCORES   = {"✅ Sí cumple": 100, "⚠️ En proceso": 50,
               "❌ No cumple": 0,   "❓ Sin evaluar": None}

SMMLV_2026 = 1_750_905

def semaforo(pct):
    if pct >= 75: return SUCCESS, "🟢 Bueno"
    if pct >= 50: return WARN,    "🟡 En proceso"
    return DANGER,  "🔴 Requiere atención"

def barra_html(pct, color):
    return (f"<div style='background:{BORDER};border-radius:20px;height:10px;width:100%'>"
            f"<div style='background:{color};width:{pct:.0f}%;height:10px;border-radius:20px;"
            f"transition:width .4s'></div></div>")

# ── FUNCIÓN PRINCIPAL ──────────────────────────────────────────────────────
def show():
    st.markdown("## 🏛️ Diagnóstico NIIF Pymes")
    st.markdown(
        f"<p style='color:{TEXT_SEC}'>Responde 10 preguntas sobre tu empresa y obtén tu nivel de cumplimiento NIIF "
        f"con un plan de acción priorizado.</p>", unsafe_allow_html=True)

    tab1, tab2 = st.tabs(["📋 Diagnóstico", "📄 Reporte"])

    # ═══════════════════════════════════════════════════════════
    # TAB 1 — DIAGNÓSTICO
    # ═══════════════════════════════════════════════════════════
    with tab1:
        # Info empresa
        col1, col2 = st.columns([2, 1])
        nombre_emp = col1.text_input("Nombre de la empresa", key="nd_nombre",
                                     placeholder="Ej: Comercializadora XYZ S.A.S.")
        grupo = col2.selectbox("Marco contable", ["Grupo 2 — NIIF Pymes", "Grupo 3 — Simplificado"],
                               key="nd_grupo")

        st.markdown("---")

        # ── Preguntas + semáforo en tiempo real ────────────────
        scores = []
        pendientes = []

        for area in AREAS:
            val = st.session_state.get(f"nd_{area['id']}", "❓ Sin evaluar")
            score = OP_SCORES[val]
            if score is not None:
                scores.append(score)
            color = OP_COLORES[val]

            with st.container():
                c_titulo, c_estado = st.columns([4, 1])

                # Barra de progreso del área
                pct_area = score if score is not None else 0
                bar_color = color if val != "❓ Sin evaluar" else BORDER

                c_titulo.markdown(
                    f"<p style='color:#E8F4FD;font-weight:600;font-size:.93rem;margin:0 0 .1rem'>"
                    f"{area['titulo']}</p>"
                    f"<p style='color:{TEXT_SEC};font-size:.8rem;margin:0 0 .4rem'>"
                    f"{area['pregunta']}</p>"
                    + barra_html(pct_area, bar_color),
                    unsafe_allow_html=True)

                nuevo_val = c_estado.selectbox(
                    "", OPCIONES,
                    index=OPCIONES.index(val),
                    key=f"nd_{area['id']}",
                    label_visibility="collapsed"
                )

                # Detalle según respuesta
                if nuevo_val != "❓ Sin evaluar":
                    if nuevo_val == "✅ Sí cumple":
                        detalle = area["si"]
                        d_color = SUCCESS
                    elif nuevo_val == "⚠️ En proceso":
                        detalle = area["proceso"]
                        d_color = WARN
                    else:
                        detalle = area["no"]
                        d_color = DANGER
                        pendientes.append(area)

                    st.markdown(
                        f"<div style='background:{d_color}11;border-left:3px solid {d_color};"
                        f"border-radius:4px;padding:.4rem .9rem;margin:.3rem 0 .6rem'>"
                        f"<p style='color:{d_color};font-size:.8rem;margin:0'>{detalle}</p></div>",
                        unsafe_allow_html=True)
                elif nuevo_val == "⚠️ En proceso":
                    pendientes.append(area)

                st.markdown("<div style='height:.2rem'></div>", unsafe_allow_html=True)

        # ── Panel de resultados ────────────────────────────────
        st.markdown("---")
        respondidas = len(scores)
        if respondidas == 0:
            st.info("👆 Responde las preguntas arriba para ver tu nivel de cumplimiento.")
        else:
            pct_global = sum(scores) / respondidas
            color_g, label_g = semaforo(pct_global)

            # Tarjeta resumen
            st.markdown(
                f"<div style='background:{BG_CARD};border:2px solid {color_g}33;"
                f"border-radius:12px;padding:1.2rem 1.5rem;text-align:center;margin-bottom:1rem'>"
                f"<p style='color:{TEXT_SEC};font-size:.82rem;margin:0 0 .2rem'>"
                f"{respondidas} de {len(AREAS)} áreas evaluadas</p>"
                f"<p style='color:{color_g};font-weight:700;font-size:2rem;margin:0'>"
                f"{pct_global:.0f}%</p>"
                f"<p style='color:{color_g};font-size:.95rem;font-weight:600;margin:.1rem 0 .5rem'>"
                f"{label_g}</p>"
                + barra_html(pct_global, color_g) +
                f"</div>",
                unsafe_allow_html=True)

            # Gráfico araña / radar
            if respondidas >= 3:
                areas_eval  = [a for a in AREAS if OP_SCORES.get(
                    st.session_state.get(f"nd_{a['id']}", "❓ Sin evaluar")) is not None]
                nombres_r   = [a["titulo"].split(" ", 1)[1][:20] for a in areas_eval]
                valores_r   = [OP_SCORES[st.session_state.get(f"nd_{a['id']}", "❓ Sin evaluar")]
                               for a in areas_eval]
                nombres_r  += [nombres_r[0]]
                valores_r  += [valores_r[0]]

                fig = go.Figure(go.Scatterpolar(
                    r=valores_r, theta=nombres_r, fill="toself",
                    fillcolor=f"{ACCENT}22", line_color=ACCENT, line_width=2,
                ))
                fig.update_layout(
                    polar=dict(
                        bgcolor=BG_CARD,
                        radialaxis=dict(visible=True, range=[0, 100],
                                        gridcolor=BORDER, tickfont_color=TEXT_SEC),
                        angularaxis=dict(gridcolor=BORDER, tickfont_color="#E8F4FD",
                                         tickfont_size=10),
                    ),
                    paper_bgcolor=BG_DARK, font_color="#E8F4FD",
                    height=340, margin=dict(l=40, r=40, t=20, b=20),
                )
                st.plotly_chart(fig, use_container_width=True)

            # Plan de acción
            gaps = [a for a in AREAS
                    if st.session_state.get(f"nd_{a['id']}", "❓ Sin evaluar")
                    in ("❌ No cumple", "⚠️ En proceso")]
            if gaps:
                st.markdown(
                    f"<p style='color:{ACCENT};font-weight:600;margin:.5rem 0 .4rem'>"
                    f"📌 Acciones prioritarias ({len(gaps)} áreas)</p>",
                    unsafe_allow_html=True)
                for a in gaps:
                    est = st.session_state.get(f"nd_{a['id']}")
                    ec  = DANGER if est == "❌ No cumple" else WARN
                    st.markdown(
                        f"<div style='background:{BG_CARD};border-left:3px solid {ec};"
                        f"border-radius:6px;padding:.6rem 1rem;margin-bottom:.4rem'>"
                        f"<p style='color:#E8F4FD;font-weight:600;font-size:.86rem;margin:0 0 .15rem'>"
                        f"{a['titulo']}</p>"
                        f"<p style='color:{TEXT_SEC};font-size:.81rem;margin:0'>{a['accion']}</p>"
                        f"</div>",
                        unsafe_allow_html=True)

            # CTA asesoría
            st.markdown(
                f"<div style='background:{BG_CARD};border-left:4px solid {ACCENT};"
                f"border-radius:8px;padding:.9rem 1.2rem;margin-top:.8rem'>"
                f"<p style='color:{ACCENT};font-weight:600;margin:0 0 .2rem'>"
                f"💼 ¿Necesitas apoyo para cerrar estas brechas?</p>"
                f"<p style='color:{TEXT_SEC};font-size:.83rem;margin:0'>"
                f"En SalazAnalytics te acompañamos con convergencia NIIF, políticas contables, "
                f"re-expresión de estados financieros, auditoría NIA y presentación a "
                f"Supersociedades, SIC y DIAN.</p></div>",
                unsafe_allow_html=True)

    # ═══════════════════════════════════════════════════════════
    # TAB 2 — REPORTE
    # ═══════════════════════════════════════════════════════════
    with tab2:
        nombre_emp = st.session_state.get("nd_nombre", "Mi Empresa")
        grupo_val  = st.session_state.get("nd_grupo",  "Grupo 2 — NIIF Pymes")
        scores_r   = [OP_SCORES[st.session_state.get(f"nd_{a['id']}", "❓ Sin evaluar")]
                      for a in AREAS
                      if OP_SCORES[st.session_state.get(f"nd_{a['id']}", "❓ Sin evaluar")] is not None]
        pct_r      = sum(scores_r) / len(scores_r) if scores_r else 0
        color_r, label_r = semaforo(pct_r)
        fecha_hoy  = datetime.now().strftime("%d de %B de %Y")

        st.markdown("### 📄 Reporte de diagnóstico")
        st.markdown(
            f"<div style='background:{BG_CARD};border:1px solid {BORDER};border-radius:12px;"
            f"padding:1.2rem 1.5rem;margin-bottom:1rem'>"
            f"<p style='color:{ACCENT};font-weight:700;font-size:1.05rem;margin:0 0 .2rem'>"
            f"🏛️ Diagnóstico NIIF — {nombre_emp or 'Mi Empresa'}</p>"
            f"<p style='color:{TEXT_SEC};font-size:.8rem;margin:0 0 .6rem'>"
            f"Generado el {fecha_hoy} · {grupo_val} · SalazAnalytics</p>"
            f"<p style='color:{color_r};font-weight:700;font-size:1.3rem;margin:0'>"
            f"{pct_r:.0f}% cumplimiento — {label_r}</p>"
            f"<p style='color:{TEXT_SEC};font-size:.81rem;margin:.3rem 0 0'>"
            f"{len(scores_r)} de {len(AREAS)} áreas evaluadas</p></div>",
            unsafe_allow_html=True)

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb  = openpyxl.Workbook()
            ws  = wb.active
            ws.title = "Diagnóstico NIIF"

            hdr_fill = PatternFill("solid", fgColor="0D1B2A")
            sub_fill = PatternFill("solid", fgColor="132030")

            ws.merge_cells("A1:E1")
            ws["A1"] = f"DIAGNÓSTICO NIIF — {nombre_emp or 'Mi Empresa'}"
            ws["A1"].font = Font(bold=True, color="00C2FF", size=13)
            ws["A1"].fill = hdr_fill
            ws["A1"].alignment = Alignment(horizontal="center")

            ws.merge_cells("A2:E2")
            ws["A2"] = f"Generado: {fecha_hoy} | {grupo_val} | Cumplimiento global: {pct_r:.0f}% — {label_r}"
            ws["A2"].font = Font(color="7B9BB5", size=10)
            ws["A2"].fill = hdr_fill
            ws["A2"].alignment = Alignment(horizontal="center")

            ws.append([])
            for h, col in zip(["Área", "Estado", "Detalle", "Acción recomendada"], range(1, 5)):
                c = ws.cell(row=4, column=col, value=h)
                c.font = Font(bold=True, color="00C2FF", size=10)
                c.fill = hdr_fill

            STATUS_MAP = {
                "✅ Sí cumple":  ("✅ Cumple",      "00FFB3"),
                "⚠️ En proceso": ("⚠️ En proceso",  "FFD93D"),
                "❌ No cumple":  ("❌ No cumple",    "FF6B6B"),
                "❓ Sin evaluar": ("— Sin evaluar",  "7B9BB5"),
            }
            for a in AREAS:
                val  = st.session_state.get(f"nd_{a['id']}", "❓ Sin evaluar")
                st_label, st_color = STATUS_MAP[val]
                if val == "✅ Sí cumple":   detalle = a["si"]
                elif val == "⚠️ En proceso": detalle = a["proceso"]
                elif val == "❌ No cumple":  detalle = a["no"]
                else:                        detalle = "No evaluado"
                row = ws.max_row + 1
                ws.cell(row=row, column=1, value=a["titulo"]).fill = sub_fill
                c2 = ws.cell(row=row, column=2, value=st_label)
                c2.font = Font(color=st_color, bold=True, size=9)
                c2.fill = sub_fill
                ws.cell(row=row, column=3, value=detalle).fill = sub_fill
                ws.cell(row=row, column=4, value=a["accion"] if val != "✅ Sí cumple" else "—").fill = sub_fill
                for col in range(1, 5):
                    ws.cell(row=row, column=col).font = ws.cell(row=row, column=col).font.copy(
                        color=st_color if col == 2 else "E8F4FD", size=9)
                    ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")

            for col, w in zip("ABCD", [35, 18, 55, 50]):
                ws.column_dimensions[col].width = w

            buf = BytesIO()
            wb.save(buf); buf.seek(0)
            st.download_button(
                "📥 Descargar reporte Excel", data=buf,
                file_name=f"Diagnostico_NIIF_{(nombre_emp or 'empresa').replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        except Exception as e:
            st.error(f"Error generando Excel: {e}")
